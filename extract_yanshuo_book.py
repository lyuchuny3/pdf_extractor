
import fitz                    #pip install pymupdf
from openpyxl import Workbook  #pip install openpyxl
import re

INFO = ['ID','NAME','STARS','TYPE','GRADE-FRENCH','GRADE-YDS','HEIGHT','BOLTS','FA','COMMENTS']
SMILE ='☺'


def get_info_page(doc):
    '''
    if the page contain "- Route Information", then return true
    '''
    def is_info_block(page):
        blocks=page.get_text("blocks")
        for block in blocks:
            if ' - Route Information' in block[-3]:
                return 1
        return 0
    page_num = doc.page_count
    real_page_idxs=[]
    for i in range(page_num):
        page=doc.load_page(i)
        if is_info_block(page):
            real_page_idxs.append(i)
    return real_page_idxs
def get_mountain_name(blocks):
    '''
    get mountain name from the line 'xxx - Route Information'
    '''
    for line in blocks:
        if 'Route Information' in line[-3]:
            mountain_name = line[-3].split('-')[0]
            return mountain_name.strip()
    return ''
def get_route_info(blocks):
    '''
    get detail route info
    - id
    - name
    - stars
    - height
    - grade(french, yds)
    - bolts
    - FA
    - comments
    '''
    todo_blocks_idx=[]
    ret = []
    for idx,block in enumerate(blocks):
        lines = block[-3].strip().split('\n')
        num = lines[0].split('.')[0]
        if not num.isdigit():
            todo_blocks_idx.append(idx)
            continue
        info ={}
        info['ID']=num
        info['STARS']=len(re.compile("☺").findall(lines[0]))*SMILE
        line0= lines[0].split('☺')[0]
        line0 = line0[line0.find('.')+1:].strip()
        name= re.compile('\D+').findall(line0)[0].strip()
        if name[-1]==',':
            name=name[:-1]
        line0=line0.replace(name,'')
        info['NAME']=name
        info['HEIGHT']=''
        info['TYPE']=''
        info['GRADE-FRENCH']=''
        info['GRADE-YDS']=''
        info['BOLTS']=''
        ################################################
        height = re.compile('\d{1,3}m').findall(line0)
        if len(height)==1:
            info['HEIGHT']=height[0]
        grade_info = re.compile('\(?\d.\d{1,2}\+?\-?[abcd]?\)?').findall(line0)
        if len(grade_info)==1:
            info['GRADE-YDS']=grade_info[0]
        type_info = re.compile('sport|traditional|top rope').findall(line0)
        if len(type_info)==1:
            info['TYPE']=type_info[0]
        grade_french_info = re.compile('\d[abcdABCD]?\+?[\s\(]').findall(line0)
        if len(grade_french_info)>=1:
            info['GRADE-FRENCH']=grade_french_info[0].strip()
        ###############################################
        # split blocks into line0, sports, descrip, FA
        FA_block=None
        Comment_block=None
        data=block[-3]
        no_line0 = data.replace(lines[0],'')

        if 'F.A.' in data:
            out =no_line0.split('F.A.')
            Comment_block=out[0].strip()
            FA_block=out[1:]
        if 'FA ' in data:
            out =no_line0.split('FA ')
            Comment_block=out[0].strip()
            FA_block=out[1:]

        if Comment_block is not None:
            '''
            get multiple info from comment part
            - bolts
            - sport
            - stars
            '''
            Comment_block=Comment_block.replace('\n','')
            bolts_info = re.compile('\d{1,2}\s*[Bb]olts').findall(Comment_block)
            if len(bolts_info)>0:
                info['BOLTS']=bolts_info[0]
                Comment_block=Comment_block.replace(bolts_info[0]+'.',"")
            sport_info = re.compile('sport').findall(Comment_block)
            if len(sport_info)>0:
                info['TYPE']=sport_info[0]
                Comment_block=Comment_block.replace(sport_info[0]+',',"")
            star_info = re.compile("☺").findall(Comment_block)
            if len(star_info)>0:
                info['STARS']=len(star_info)*SMILE
                Comment_block=Comment_block.replace(len(star_info)*SMILE,"")
            Comment_block=Comment_block.strip()
        info['FA']=''
        if FA_block:
            info['FA']=';'.join([ss.strip() for ss in FA_block])
        
        info['COMMENTS']=Comment_block
        ret.append(info)
    return ret
def get_all_mountain_info(infile_name,outfile_name):
    '''
    - open the pdf
    - get info pages
    - get route info from page
    - write info to excel
    '''
    doc = fitz.open(infile_name)
    #init wb
    wb = Workbook()
    del wb['Sheet']
    
    real_pages=get_info_page(doc)
    for page_i in real_pages:
        page = doc.load_page(page_i)
        blocks=page.get_text("blocks")
        MOUNTAIN_NAME = get_mountain_name(blocks)
        route_info = get_route_info(blocks)
        
        # save into excel            
        try:
            sh = wb[MOUNTAIN_NAME]
        except:
            sh = wb.create_sheet(MOUNTAIN_NAME)
            sh.title = MOUNTAIN_NAME
            for idx,item in enumerate(INFO):
                sh.cell(1,idx+1).value=item
        num_rows = sh.max_row
        for idx,info in enumerate(route_info):
            for j,item in enumerate(INFO):
                sh.cell(num_rows+idx+1,j+1).value = info[item]
    wb.save(filename=outfile_name)
    
    
get_all_mountain_info("yanshuo_routebook_2010.pdf","yanshuo.xlsx")

# # for debug:
# # - get page number 
# doc = fitz.open("a.pdf")
# read_pages = get_info_page(doc)
# page_num = 18
# blocks=doc.load_page(page_num).get_text("blocks")
# lines = blocks[14][-3].strip().split('\n')
# height = re.compile('\d{1,3}m').findall(lines[0])
# print(height[0])
